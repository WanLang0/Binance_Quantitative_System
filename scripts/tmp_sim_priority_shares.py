# -*- coding: utf-8 -*-
"""模拟「量化优先匹配 · 少份额」下的实际成交胜率 vs 回测全信号胜率

方法：读前20三档Excel逐笔明细 → 每档把 20 个币的全部开/平仓事件按时间全局排序
→ 用 N 份槽位模拟先到先得（有空槽才成交，平仓释放槽位）→ 统计成交子集胜率/收益。
多轮打乱同时段信号的先后顺序（同K线内先后其实随机），看胜率分布的波动。
"""
import os, sys, io, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import datetime

import numpy as np
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'results', 'top20_macd_vol_trades_detail_2024_2026.xlsx')
SHARES = [2, 4, 6, 10]
ROUNDS = 200  # 打乱轮数


def load_sheet(ws):
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {h: i for i, h in enumerate(header)}
    trades = []
    for r in rows:
        if r[col['开仓时间']] is None:
            continue
        t0, t1 = r[col['开仓时间']], r[col['平仓时间']]
        if isinstance(t0, str):
            t0 = datetime.fromisoformat(t0)
        if isinstance(t1, str):
            t1 = datetime.fromisoformat(t1)
        trades.append((t0, t1, float(r[col['收益率(%)']])))
    return trades


def simulate_shares(trades, n_shares, rng):
    """先到先得：开仓事件按时间排序，有空槽成交否则跳过；平仓释放槽。
    同一时刻多信号竞争时顺序随机（打乱键加微小抖动）。"""
    evs = []
    for i, (t0, t1, ret) in enumerate(trades):
        jitter = rng.random() / 86400.0  # 同秒内随机先后
        evs.append((t0.timestamp() + jitter, 0, i))    # 0=开仓
        evs.append((t1.timestamp() + jitter, 1, i))    # 1=平仓
    evs.sort()
    busy = 0
    taken = [False] * len(trades)
    for ts, kind, i in evs:
        if kind == 0:
            if busy < n_shares:
                busy += 1
                taken[i] = True
        else:
            if taken[i]:
                busy -= 1
    rets = [trades[i][2] for i in range(len(trades)) if taken[i]]
    return rets


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    rng = random.Random(42)
    for name in wb.sheetnames:
        if name == '汇总':
            continue
        ws = wb[name]
        trades = load_sheet(ws)
        all_rets = [t[2] for t in trades]
        base_wr = np.mean([r > 0 for r in all_rets]) * 100
        base_avg = np.mean(all_rets)
        print('\n' + '=' * 100)
        print('【%s】 全信号 %d 笔 | 胜率 %.1f%% | 笔均 %+.2f%%' % (
            name, len(trades), base_wr, base_avg))

        for n_sh in SHARES:
            wrs, avgs, cnts = [], [], []
            for _ in range(ROUNDS):
                rets = simulate_shares(trades, n_sh, rng)
                if not rets:
                    continue
                wrs.append(np.mean([r > 0 for r in rets]) * 100)
                avgs.append(np.mean(rets))
                cnts.append(len(rets))
            wrs, avgs, cnts = map(np.array, (wrs, avgs, cnts))
            print('  %2d份: 成交%5.0f笔(全集%.0f%%) 胜率 均%.1f%% | 最好%.1f%% 最差%.1f%% | '
                  '胜率<40%%的轮数占比%.0f%% | 笔均 %+.2f%%' % (
                      n_sh, cnts.mean(), cnts.mean() / len(trades) * 100,
                      wrs.mean(), wrs.max(), wrs.min(),
                      (wrs < 40).mean() * 100, avgs.mean()))


if __name__ == '__main__':
    main()

# -*- coding: utf-8 -*-
"""三种推荐配置（虚拟币 macd+背离+量能）交易频次与开仓时间分布分析
数据源：scripts/results/top20_macd_vol_trades_detail_2024_2026.xlsx（逐笔明细）
时间口径：币安K线为 UTC；北京时间 = UTC+8
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import Counter
from datetime import datetime, timedelta

import openpyxl

XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'results',
                    'top20_macd_vol_trades_detail_2024_2026.xlsx')

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
print('sheets:', wb.sheetnames)

for name in wb.sheetnames:
    if name == '汇总':
        continue
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    # 定位列
    col = {h: i for i, h in enumerate(header)}
    i_open = col['开仓时间']
    i_sym = col['币种']
    i_dir = col.get('方向')
    i_year = col.get('年份')
    i_close = col.get('平仓时间')

    total = 0
    hour_utc = Counter()      # UTC小时
    weekday = Counter()       # UTC星期(0=周一)
    year_cnt = Counter()
    sym_cnt = Counter()
    dir_cnt = Counter()
    hold_hours = []
    open_pts = Counter()      # 开仓在周期内的分钟点（15m→0/15/30/45；1h→0）

    for r in rows:
        if r[i_open] is None:
            continue
        t = r[i_open]
        if isinstance(t, str):
            t = datetime.fromisoformat(t)
        if not isinstance(t, datetime):
            continue
        total += 1
        hour_utc[t.hour] += 1
        weekday[t.weekday()] += 1
        year_cnt[str(t.year)] += 1
        sym_cnt[r[i_sym]] += 1
        if i_dir is not None and r[i_dir]:
            dir_cnt[r[i_dir]] += 1
        open_pts[t.minute] += 1
        c = r[i_close] if i_close is not None else None
        if isinstance(c, str):
            c = datetime.fromisoformat(c)
        if isinstance(c, datetime):
            hold_hours.append((c - t).total_seconds() / 3600.0)

    print('\n' + '=' * 96)
    print('【%s】 总笔数 %d' % (name, total))
    print('  分年:', dict(sorted(year_cnt.items())), '| 20币均 %.0f 笔/币·3年' % (total / 20))
    print('  方向:', dict(dir_cnt) if dir_cnt else '—')
    if hold_hours:
        hs = sorted(hold_hours)
        print('  持仓时长: 中位 %.1fh  均 %.1fh  P90 %.1fh  最长 %.0fh' % (
            hs[len(hs) // 2], sum(hs) / len(hs), hs[int(len(hs) * 0.9)], hs[-1]))

    # 小时分布（UTC 与 北京时间）
    n = max(total, 1)
    print('  开仓小时分布（按北京时间，括号内=UTC）:')
    line1 = []; line2 = []
    for bj in range(24):
        utc = (bj - 8) % 24
        p = hour_utc.get(utc, 0) / n * 100
        bar = '#' * int(p / 1.5)
        line1.append('%02d时(%02d) %5.1f%% %s' % (bj, utc, p, bar))
    for i in range(0, 24, 2):
        print('    ' + ' | '.join(line1[i:i + 2]))

    top3 = hour_utc.most_common(3)
    top3_pct = sum(c for _, c in top3) / n * 100
    print('  最活跃3小时(UTC): %s 合计 %.1f%%（均匀基线12.5%%）' % (
        ', '.join('%02d时 %.1f%%' % (h, c / n * 100) for h, c in top3), top3_pct))
    lo = hour_utc.most_common()[-3:]
    print('  最冷清3小时(UTC): %s' % ', '.join('%02d时 %.1f%%' % (h, c / n * 100) for h, c in lo))
    # 美股时段（北京21:00-次日4:00=UTC 13-20）占比
    us_pct = sum(hour_utc.get(u, 0) for u in range(13, 21)) / n * 100
    print('  美股交易时段(UTC13-20/北京21-04)占比: %.1f%%' % us_pct)

    wd = ['一', '二', '三', '四', '五', '六', '日']
    print('  星期分布(UTC):', ' '.join('%s%.1f%%' % (wd[i], weekday.get(i, 0) / n * 100) for i in range(7)))

    if open_pts:
        pts = sorted(open_pts.items())
        print('  开仓分钟点分布:', ', '.join('%02d分:%d' % (m, c) for m, c in pts if c > total * 0.01))

    top_sym = sym_cnt.most_common(5)
    print('  交易最多5币:', ', '.join('%s %d' % (s, c) for s, c in top_sym))
    bot_sym = sym_cnt.most_common()[-3:]
    print('  交易最少3币:', ', '.join('%s %d' % (s, c) for s, c in bot_sym))

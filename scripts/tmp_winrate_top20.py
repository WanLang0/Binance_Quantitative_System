# -*- coding: utf-8 -*-
"""虚拟币三档配置单笔胜率统计（前20：读Excel逐笔明细；前40：缓存K线重算逐笔）"""
import os, sys, io, json, warnings
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
warnings.filterwarnings('ignore')
from collections import Counter
from datetime import datetime, timezone

import numpy as np
import pandas as pd
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'results', 'top20_macd_vol_trades_detail_2024_2026.xlsx')

print('========== 前20（市值1-20，Excel逐笔明细） ==========')
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
for name in wb.sheetnames:
    if name == '汇总':
        continue
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {h: i for i, h in enumerate(header)}
    i_ret, i_dir, i_year, i_sym = col['收益率(%)'], col['方向'], col['年份'], col['币种']
    i_pnl = col.get('净盈亏')

    n = win = 0
    wins, losses = [], []
    per_year = {}
    per_dir = Counter(); per_dir_win = Counter()
    for r in rows:
        if r[i_ret] is None:
            continue
        ret = float(r[i_ret]); n += 1
        y = str(r[i_year]); d = str(r[i_dir])
        py = per_year.setdefault(y, [0, 0]); py[0] += 1
        if ret > 0:
            win += 1; wins.append(ret); py[1] += 1
            per_dir_win[d] += 1
        else:
            losses.append(ret)
        per_dir[d] += 1
    if n == 0:
        continue
    aw = np.mean(wins) if wins else 0
    al = np.mean(losses) if losses else 0
    print('\n【%s】 总%d笔' % (name, n))
    print('  总胜率: %d/%d = %.1f%%' % (win, n, win / n * 100))
    print('  均盈利单 %+.2f%% / 均亏损单 %+.2f%% / 盈亏比 %.2f' % (
        aw, al, (aw / abs(al)) if al else float('inf')))
    print('  最大单笔盈利 %+.1f%% / 最大单笔亏损 %.1f%%' % (max(wins) if wins else 0, min(losses) if losses else 0))
    ys = '  '.join('%s:%.1f%%(%d/%d)' % (y, w[1] / w[0] * 100, w[1], w[0]) for y, w in sorted(per_year.items()))
    print('  分年胜率:', ys)
    ds = '  '.join('%s:%.1f%%(%d/%d)' % (d, per_dir_win[d] / per_dir[d] * 100, per_dir_win[d], per_dir[d])
                   for d in per_dir)
    print('  分方向胜率:', ds)
